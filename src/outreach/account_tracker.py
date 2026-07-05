"""Priority company account tracker and scorer.

Scoring design decisions: docs/relationship_engine.md → Scoring Philosophy
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from outreach.tracking import (
    ContactRecord,
    OpportunityRecord,
    OrganizationRecord,
    OutreachWorkbook,
)

# ---------------------------------------------------------------------------
# Profile fit domains (earned credentials only — see relationship_engine.md)
# ---------------------------------------------------------------------------

DOMAIN_TAGS: dict[str, int] = {
    # AI/ML products — FlairX, L'Oréal project
    "artificial-intelligence": 5,
    "ai": 5,
    "machine-learning": 5,
    "generative-ai": 5,
    "applied-ai": 5,
    "llm": 5,
    "large-language-model": 5,
    "agent": 4,
    "conversational-ai": 4,
    "nlp": 4,
    # Data infrastructure / data platforms — Hevo Data
    "data": 5,
    "data-infrastructure": 5,
    "data-platform": 5,
    "data-pipeline": 5,
    "etl": 5,
    "analytics": 4,
    "business-intelligence": 4,
    "warehousing": 4,
    # API / integration platforms — Hevo connectors
    "integration": 4,
    "api": 4,
    "connectivity": 4,
    # Observability / monitoring — Hevo 2.0 AI monitoring
    "observability": 5,
    "monitoring": 4,
    "devops": 3,
    "platform-engineering": 4,
    # Developer tools / DevEx
    "developer-tools": 5,
    "developer-experience": 5,
    "devex": 5,
    "developer-platform": 5,
    "infrastructure": 4,
    # Hiring tech / workflow automation — FlairX + ResumeGenerator
    "hiring": 5,
    "recruiting": 5,
    "hr-tech": 5,
    "talent": 4,
    "workflow-automation": 4,
    "automation": 3,
    # Consumer marketplace / logistics — Gojek
    "marketplace": 4,
    "logistics": 4,
    "mobility": 4,
    "transportation": 4,
    "delivery": 3,
    "gig-economy": 3,
    # Supply chain / ops tech — Gojek fleet
    "supply-chain": 3,
    "fleet": 3,
    "operations": 3,
    # Enterprise SaaS with product culture — Notion, Rippling, Ramp, Airtable type
    "productivity": 3,
    "collaboration": 3,
    "saas": 2,
    # FinTech / billing — Intuit
    "fintech": 4,
    "payments": 4,
    "financial-technology": 4,
    "billing": 4,
    "banking": 3,
    # Healthcare IT — Optum
    "healthcare": 3,
    "health-tech": 3,
    "healthtech": 3,
    "digital-health": 3,
    "medtech": 3,
    # AI agents / autonomous workflow — ResumeGenerator, Outreach
    "agentic": 5,
    "autonomous": 4,
    "copilot": 4,
}

PROFILE_FIT_CAP = 25
LINKEDIN_WAVE_SIZE = 8
MIN_MAPPED_CONTACTS = 3
BRAND_SCORE_MAX = 12
PITCH_SCORE_MAX = 10
MANUAL_PRIORITY_TAGS = {"priority", "core", "relationship", "target", "dream", "track-2", "tier-a"}
RELATIONSHIP_TIER_A_TOTAL = 32
TIER_B_TOTAL = 50
TIER_A_MIN_SCORE = 30
TIER_B_MIN_SCORE = 18
TIER_A_TRACK_QUOTAS = {
    "Startup / Founder-Led": 20,
    "Growth / Mid-Market": 12,
}
LARGE_COMPANY_L1_TOTAL = 20
LARGE_COMPANY_L2_TOTAL = 40

# ---------------------------------------------------------------------------
# Role fit patterns
# ---------------------------------------------------------------------------

ROLE_FIT_PATTERNS: list[tuple[int, list[str]]] = [
    (25, ["pm intern", "product intern", "product manager intern", "apm intern",
          "associate product manager intern", "mba product manager intern",
          "mba pm intern"]),
    (22, ["apm", "associate product manager", "product ops intern",
          "product operations intern", "ai pm", "technical pm intern"]),
    (18, ["technical pm", "technical product manager", "product manager",
          "founder's associate intern", "strategy intern", "chief of staff intern",
          "growth product intern", "platform product intern"]),
    (12, ["founder's associate", "chief of staff", "strategy", "product ops",
          "product operations", "business operations", "bizops", "growth intern",
          "program manager intern", "tpm intern"]),
    (6,  ["software engineer intern", "engineering intern", "data science intern",
          "data analyst intern"]),
]

# ---------------------------------------------------------------------------
# Reachability signals (Akshat-specific only)
# ---------------------------------------------------------------------------

SHARED_EMPLOYERS = ["intuit", "gojek", "hevo", "optum"]
LA_KEYWORDS = ["los angeles", " la,", "santa monica", "culver city", "el segundo",
               "manhattan beach", "playa vista", "west la", "westwood",
               "burbank", "glendale", "pasadena", "long beach", "irvine"]
REMOTE_KEYWORDS = ["remote", "work from home", "wfh", "distributed"]

# ---------------------------------------------------------------------------
# Account stages
# ---------------------------------------------------------------------------

INVITED_STATUSES = {"Invited", "Invite error", "Connected", "Warm", "Followed up", "Replied"}
ACCEPTED_STATUSES = {"Connected", "Warm", "Followed up", "Replied"}
REPLIED_STATUSES = {"Replied"}


# ---------------------------------------------------------------------------
# Data class
# ---------------------------------------------------------------------------

@dataclass
class AccountRow:
    organization_id: str
    company: str
    org_type: str
    city: str
    website: str
    fit_score: int = 0
    account_score: int = 0
    tier: str = "C"
    why_fit: str = ""
    target_role: str = ""
    hiring_signal: str = "No open roles"
    people_mapped: int = 0
    email_contacts: int = 0
    invites_sent: int = 0
    accepted: int = 0
    replies: int = 0
    coffee_chats: int = 0
    advocates: int = 0
    account_stage: str = "unqualified"
    next_action: str = ""
    next_due_date: str = ""
    target_lists: str = ""
    team_size: Optional[int] = None
    tags: str = ""
    score_profile_fit: int = 0
    score_role_fit: int = 0
    score_team_gate: int = 0
    score_reachability: int = 0
    score_hiring: int = 0
    score_relationship: int = 0
    score_relationship_momentum: int = 0
    score_brand: int = 0
    score_pitch_strength: int = 0
    score_account_hiring: int = 0
    data_quality_flags: str = ""
    campaign_action: str = ""
    campaign_channel: str = ""
    campaign_priority: int = 0
    daily_action_priority: int = 0
    campaign_reason: str = ""
    lane_1_policy: str = ""
    account_track: str = ""


@dataclass
class CampaignPlanRow:
    company: str
    tier: str
    fit_score: int
    account_score: int
    account_stage: str
    campaign_action: str
    campaign_channel: str
    account_track: str
    campaign_priority: int
    daily_action_priority: int
    campaign_reason: str
    lane_1_policy: str
    why_fit: str
    people_mapped: int
    email_contacts: int
    invites_sent: int
    accepted: int
    replies: int
    target_role: str
    next_due_date: str
    organization_id: str


@dataclass(frozen=True)
class DailyPlanBudget:
    max_total_actions: int = 24
    max_companies: int = 18
    max_linkedin_invites: int = 12
    max_linkedin_followups: int = 8
    max_company_mapping: int = 5
    max_email_research: int = 5
    max_context_enrichment: int = 8
    max_email_drafts: int = 0


@dataclass
class DailyPlanItem:
    company: str
    tier: str
    campaign_action: str
    campaign_channel: str
    daily_action_priority: int
    account_score: int
    phase: str = ""
    phase_order: int = 0
    can_parallelize: bool = False
    expected_linkedin_invites: int = 0
    expected_linkedin_followups: int = 0
    expected_company_mapping: int = 0
    expected_email_research: int = 0
    expected_context_enrichment: int = 0
    expected_email_drafts: int = 0
    reason: str = ""
    skip_reason: str = ""
    organization_id: str = ""


# ---------------------------------------------------------------------------
# Scoring helpers
# ---------------------------------------------------------------------------

def _parse_notes(notes: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for part in notes.split("|"):
        part = part.strip()
        if "=" in part:
            k, _, v = part.partition("=")
            result[k.strip()] = v.strip()
    return result


def _parse_team_size(value: str) -> Optional[int]:
    match = re.search(r"\d[\d,]*", value or "")
    if not match:
        return None
    return int(match.group(0).replace(",", ""))


def _normalize_tag(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def _split_normalized_tags(value: str) -> set[str]:
    return {
        tag
        for tag in (_normalize_tag(part) for part in re.split(r"[,;|]", value or ""))
        if tag
    }


def _mentions_domain_term(text: str, term: str) -> bool:
    normalized = _normalize_tag(term)
    if not normalized:
        return False
    parts = [re.escape(part) for part in normalized.split("-") if part]
    if not parts:
        return False
    pattern = r"(?<![a-z0-9])" + r"[\s\-+/&]+".join(parts) + r"(?![a-z0-9])"
    return re.search(pattern, text.lower()) is not None


def _text_mentions_company(text: str, company: str) -> bool:
    text_lower = text.lower()
    company_lower = company.lower().strip()
    if not company_lower:
        return False
    if len(company_lower) <= 6:
        return re.search(rf"(?<![a-z0-9]){re.escape(company_lower)}(?![a-z0-9])", text_lower) is not None
    return company_lower in text_lower


def _company_relevant_contacts(org: OrganizationRecord, contacts: list[ContactRecord]) -> list[ContactRecord]:
    relevant = [
        contact
        for contact in contacts
        if _text_mentions_company(" ".join([contact.title, contact.notes, contact.source_url]), org.name)
    ]
    # If no contact carries enough text to prove company fit, fall back to the
    # org assignment so manually curated company contacts still count.
    return relevant or contacts


def _score_profile_fit(org: OrganizationRecord) -> tuple[int, list[str], str, Optional[int]]:
    parsed = _parse_notes(org.notes)
    tags_raw = parsed.get("tags", "").lower()
    description = parsed.get("description", "").lower()
    team_size = _parse_team_size(parsed.get("team_size", ""))
    normalized_tags = _split_normalized_tags(tags_raw)

    score = 0
    matched: list[str] = []
    for tag, pts in DOMAIN_TAGS.items():
        normalized_tag = _normalize_tag(tag)
        if normalized_tag in normalized_tags or _mentions_domain_term(description, tag):
            score += pts
            matched.append(tag)

    return min(score, PROFILE_FIT_CAP), matched, parsed.get("tags", ""), team_size


def _opportunity_text(opp: OpportunityRecord) -> str:
    return " ".join(
        [
            opp.title,
            str(opp.opportunity_type),
            opp.target_lists,
            opp.location,
            opp.source_url,
            opp.notes,
        ]
    ).lower()


def _opportunity_role_text(opp: OpportunityRecord) -> str:
    return " ".join([opp.title, opp.target_lists, opp.notes]).lower()


def _is_internship_role(opp: OpportunityRecord) -> bool:
    text = _opportunity_text(opp)
    return opp.opportunity_type == "internship" or "intern" in text


def _is_full_time_path_role(opp: OpportunityRecord) -> bool:
    text = _opportunity_text(opp)
    if _is_internship_role(opp):
        return False
    if opp.opportunity_type == "full_time":
        return True
    return any(
        pattern in text
        for pattern in (
            "full-time",
            "full time",
            "new grad",
            "new-grad",
            "early career",
            "rotational",
            "associate product manager",
            "technical product manager",
            "product manager",
            "product owner",
            "product operations",
            "product ops",
            "apm",
        )
    )


def _is_fall_or_current_internship(opp: OpportunityRecord) -> bool:
    if not _is_internship_role(opp):
        return False
    text = _opportunity_role_text(opp)
    return re.search(
        r"(?<![a-z0-9])(?:fall|autumn|spring|winter|co[-\s]?op|off[-\s]?cycle)(?![a-z0-9])",
        text,
    ) is not None


def _is_remote_opportunity(opp: OpportunityRecord) -> bool:
    return any(keyword in _opportunity_text(opp) for keyword in REMOTE_KEYWORDS)


def _is_la_compatible_opportunity(opp: OpportunityRecord) -> bool:
    text = _opportunity_text(opp)
    return any(keyword in text for keyword in LA_KEYWORDS)


def _has_known_location(opp: OpportunityRecord) -> bool:
    text = " ".join([opp.location, opp.notes]).lower()
    return bool(opp.location.strip()) or "location=" in text


def _is_fall_location_compatible(opp: OpportunityRecord) -> bool:
    if not _is_fall_or_current_internship(opp):
        return False
    return _is_remote_opportunity(opp) or _is_la_compatible_opportunity(opp)


def _is_fall_location_incompatible(opp: OpportunityRecord) -> bool:
    if not _is_fall_or_current_internship(opp):
        return False
    return _has_known_location(opp) and not _is_fall_location_compatible(opp)


def _is_summer_internship(opp: OpportunityRecord) -> bool:
    return _is_internship_role(opp) and "summer" in _opportunity_text(opp)


def _score_role_fit(opps: list[OpportunityRecord]) -> tuple[int, str, str]:
    if not opps:
        return 0, "", "No open roles"
    best_score = 0
    best_title = ""
    for opp in opps:
        title_lower = opp.title.lower()
        for pts, patterns in ROLE_FIT_PATTERNS:
            if any(p in title_lower for p in patterns):
                if pts > best_score:
                    best_score = pts
                    best_title = opp.title
                break
    if best_score == 0 and opps:
        best_score = 3
        best_title = opps[0].title
    ft_count = sum(1 for o in opps if _is_full_time_path_role(o))
    fall_count = sum(1 for o in opps if _is_fall_location_compatible(o))
    fall_unknown_count = sum(
        1
        for o in opps
        if _is_fall_or_current_internship(o)
        and not _is_fall_location_compatible(o)
        and not _is_fall_location_incompatible(o)
    )
    fall_incompatible_count = sum(1 for o in opps if _is_fall_location_incompatible(o))
    summer_count = sum(1 for o in opps if _is_summer_internship(o))
    intern_count = sum(1 for o in opps if _is_internship_role(o))
    generic_intern_count = max(0, intern_count - fall_count - fall_unknown_count - fall_incompatible_count - summer_count)
    if ft_count >= 2:
        hiring_signal = f"{ft_count} FT/product-path roles open"
    elif ft_count == 1:
        hiring_signal = "FT/product-path role open"
    elif fall_count >= 2:
        hiring_signal = f"{fall_count} fall/co-op internships, LA/remote"
    elif fall_count == 1:
        hiring_signal = "Fall/co-op internship, LA/remote"
    elif fall_unknown_count >= 2:
        hiring_signal = f"{fall_unknown_count} fall/co-op internships, location unknown"
    elif fall_unknown_count == 1:
        hiring_signal = "Fall/co-op internship, location unknown"
    elif fall_incompatible_count >= 2:
        hiring_signal = f"{fall_incompatible_count} fall/co-op internships outside LA/remote"
    elif fall_incompatible_count == 1:
        hiring_signal = "Fall/co-op internship outside LA/remote"
    elif generic_intern_count >= 2:
        hiring_signal = f"{generic_intern_count} internships discovered"
    elif generic_intern_count == 1:
        hiring_signal = "Internship discovered"
    elif summer_count >= 2:
        hiring_signal = f"{summer_count} summer internships discovered"
    elif summer_count == 1:
        hiring_signal = "Summer internship discovered"
    elif opps:
        hiring_signal = f"{len(opps)} role(s) discovered"
    else:
        hiring_signal = "No open roles"
    return min(best_score, 25), best_title, hiring_signal


def _score_team_gate(team_size: Optional[int]) -> int:
    """Team size as a company maturity gate, not reachability."""
    # TODO(data): team_size comes from org notes (parsed from YC/BuiltIn scrape).
    # Companies imported from jobs.xlsx have no team_size → returns 0 (no adjustment).
    # Enrich via Crunchbase or LinkedIn company page for better coverage.
    if team_size is None:
        return 0  # unknown: no adjustment
    if team_size < 10:
        return -10  # too small: no real PM structure
    if team_size < 15:
        return -5   # marginal
    # 15–200: sweet spot, no adjustment
    return 0


def _data_quality_flags(
    *,
    org: OrganizationRecord,
    profile_fit: int,
    team_size: Optional[int],
    opportunities: list[OpportunityRecord],
) -> list[str]:
    flags: list[str] = []
    if profile_fit == 0 and "imported from resumegenerator" in org.notes.lower():
        flags.append("needs_domain_enrichment")
    if team_size is None and "team_size=" in org.notes.lower():
        flags.append("team_size_unparsed")
    if not org.website and not org.linkedin_url:
        flags.append("missing_company_url")
    if opportunities and profile_fit == 0:
        flags.append("role_without_domain_context")
    if _parse_notes(org.notes).get("context_confidence") == "inferred_from_job":
        flags.append("context_inferred_from_job")
    return flags


def _score_brand(org: OrganizationRecord) -> tuple[int, str]:
    parsed = _parse_notes(org.notes)
    text = " ".join([org.name, org.target_lists, org.notes]).lower()
    target_tags = _split_normalized_tags(org.target_lists)
    prestige_tags = _split_normalized_tags(parsed.get("prestige_signals", ""))
    top_tier = {
        "stripe",
        "scale ai",
        "figma",
        "databricks",
        "rippling",
        "ramp",
        "openai",
        "anthropic",
        "notion",
        "airtable",
        "cursor",
        "vercel",
        "coinbase",
    }
    strong = {
        "zoom",
        "typeface",
        "deel",
        "faire",
        "checkr",
        "motive",
        "webflow",
        "pagerduty",
        "amplitude",
        "monte carlo",
        "klaviyo",
        "airwallex",
        "celonis",
        "navan",
        "maven",
    }
    company = org.name.lower().strip()
    manual_score = 0
    manual_label = ""
    if target_tags.intersection(MANUAL_PRIORITY_TAGS) or any(
        _mentions_domain_term(text, token)
        for token in ("manual priority", "tier a", "dream company", "priority account")
    ):
        manual_score = 12 if ("dream" in target_tags or "tier-a" in target_tags) else 10
        manual_label = "manual priority"

    brand_score = 0
    brand_label = ""
    if company in top_tier:
        brand_score = 12
        brand_label = "top-tier brand"
    elif company in strong:
        brand_score = 8
        brand_label = "strong brand"
    elif target_tags.intersection({"yc", "y-combinator"}) or _mentions_domain_term(text, "y combinator"):
        brand_score = 5
        brand_label = "YC signal"
    elif target_tags.intersection({"built-in", "builtin"}) or any(
        _mentions_domain_term(text, token)
        for token in ("built in", "growth stage", "series a", "series b", "series c")
    ):
        brand_score = 3
        brand_label = "recognizable company signal"

    top_backer_signals = {
        "sequoia-backed",
        "a16z-backed",
        "gv-backed",
        "accel-backed",
        "index-backed",
        "khosla-backed",
        "benchmark-backed",
        "founders-fund-backed",
    }
    prestige_score = 0
    prestige_label = ""
    if prestige_tags.intersection(top_backer_signals):
        prestige_score = 10
        prestige_label = "top investor signal"
    elif "series-c-plus" in prestige_tags:
        prestige_score = 8
        prestige_label = "growth funding signal"
    elif prestige_tags.intersection({"series-a", "series-b"}):
        prestige_score = 6
        prestige_label = "venture funding signal"
    elif "yc-backed" in prestige_tags:
        prestige_score = 5
        prestige_label = "YC signal"
    elif prestige_tags.intersection({"techcrunch-covered", "crunchbase-profile", "venture-backed", "seed-funded"}):
        prestige_score = 3
        prestige_label = "external prestige signal"

    if manual_score >= brand_score and manual_score >= prestige_score:
        return min(BRAND_SCORE_MAX, manual_score), manual_label
    if prestige_score >= brand_score:
        return min(BRAND_SCORE_MAX, prestige_score), prestige_label
    return min(BRAND_SCORE_MAX, brand_score), brand_label


def _score_pitch_strength(
    *,
    org: OrganizationRecord,
    profile_fit: int,
    reachability: int,
    team_size: Optional[int],
    data_quality_flags: list[str],
) -> tuple[int, str]:
    parsed = _parse_notes(org.notes)
    target_tags = _split_normalized_tags(org.target_lists)
    has_explicit_story_fit = bool(
        parsed.get("why_this_company")
        or parsed.get("story_fit_reason")
        or parsed.get("story_angle")
        or parsed.get("profile_evidence")
    )
    if (
        "needs_domain_enrichment" in data_quality_flags
        or "role_without_domain_context" in data_quality_flags
    ) and not has_explicit_story_fit:
        return 0, "needs domain context"
    score = 0
    reasons: list[str] = []
    if has_explicit_story_fit:
        if parsed.get("why_this_company") or parsed.get("story_fit_reason"):
            score += 5
            reasons.append("explicit story-fit pitch")
        if parsed.get("profile_evidence"):
            score += 3
            reasons.append("profile evidence")
        if parsed.get("story_angle"):
            score += 2
            reasons.append("story angle")
        if "story-fit" in target_tags:
            score += 1
    if profile_fit >= 20:
        score += 5
        reasons.append("strong profile story")
    elif profile_fit >= 12:
        score += 3
        reasons.append("usable profile story")
    if reachability >= 5:
        score += 3
        reasons.append("warm path")
    elif reachability > 0:
        score += 1
        reasons.append("some reachability")
    if team_size is not None:
        if 15 <= team_size <= 1000:
            score += 2
            reasons.append("team size supports targeted relationship work")
        elif team_size < 10:
            score -= 2
            reasons.append("very small team")
    if "context_inferred_from_job" in data_quality_flags:
        score = min(score, 3)
        reasons.append("job-inferred context")
    return max(0, min(PITCH_SCORE_MAX, score)), ", ".join(reasons)


def _score_account_hiring(role_score: int, hiring_score: int) -> int:
    if hiring_score >= 18 and role_score >= 18:
        return 6
    if hiring_score >= 12 and role_score >= 12:
        return 4
    if hiring_score >= 5:
        return 2
    return 0


def _score_account_campaign(
    *,
    profile_fit: int,
    reachability: int,
    brand: int,
    relationship: int,
    account_hiring: int,
    pitch_strength: int,
    team_gate: int,
    data_quality_flags: list[str],
) -> int:
    score = profile_fit + reachability + brand + relationship + account_hiring + pitch_strength + team_gate
    if "needs_domain_enrichment" in data_quality_flags:
        score -= 8
    elif "role_without_domain_context" in data_quality_flags:
        score -= 5
    if "context_inferred_from_job" in data_quality_flags:
        score -= 4
    return max(0, score)


def _score_reachability(
    org: OrganizationRecord,
    contacts: list[ContactRecord],
) -> tuple[int, list[str]]:
    """Akshat-specific network advantages only. Max 12 pts."""
    score = 0
    signals: list[str] = []
    notes_lower = org.notes.lower()
    city_lower = org.city.lower()

    # USC / Marshall path
    has_usc = any(
        "usc" in c.notes.lower() or "marshall" in c.notes.lower()
        for c in contacts
    )
    if has_usc:
        score += 5
        signals.append("USC path")

    # Indian / Delhi background in contacts
    has_india = any(
        "india" in c.notes.lower() or "indian" in c.notes.lower() or "delhi" in c.notes.lower()
        for c in contacts
    )
    if has_india:
        score += 4
        signals.append("India/Delhi path")

    # Shared past employer (founder or exec from Intuit/Gojek/Hevo/Optum)
    has_shared_employer = any(emp in notes_lower for emp in SHARED_EMPLOYERS)
    if not has_shared_employer:
        has_shared_employer = any(
            any(emp in c.notes.lower() for emp in SHARED_EMPLOYERS)
            for c in contacts
        )
    if has_shared_employer:
        score += 4
        signals.append("shared employer")

    # LA location
    if any(kw in city_lower for kw in LA_KEYWORDS):
        score += 3
        signals.append("LA location")

    # 2nd-degree connection density (≥2 2nd-degree contacts)
    # TODO(data): 2nd-degree status is captured by the LinkedIn Playwright pipeline
    # as `triggers=2nd Degree,...` in contact notes. This is populated automatically
    # on each daily run. No manual enrichment needed — just keep the pipeline running.
    second_deg_count = sum(
        1 for c in contacts
        if "2nd degree" in c.notes.lower() or "2nd" in c.notes.lower()
    )
    if second_deg_count >= 2:
        score += 3
        signals.append(f"{second_deg_count}× 2nd-degree")

    return min(score, 12), signals


def _score_hiring(opps: list[OpportunityRecord]) -> tuple[int, list[str]]:
    # Track 2 is account-first. Roles are timing/path signals, not the ranking
    # center. Summer internship listings are weak now; FT/new-grad/APM/fall/co-op
    # roles are more relevant for the current relationship campaign.
    if not opps:
        return 0, []
    if any(_is_full_time_path_role(o) for o in opps):
        return 18, ["FT/product path listed"]
    if any(o.opportunity_type == "full_time" for o in opps):
        return 14, ["FT role listed"]
    if any(_is_fall_location_compatible(o) for o in opps):
        return 12, ["LA/remote fall/co-op internship listed"]
    if any(
        _is_fall_or_current_internship(o)
        and not _is_fall_location_compatible(o)
        and not _is_fall_location_incompatible(o)
        for o in opps
    ):
        return 6, ["fall/co-op internship needs location check"]
    if any(_is_fall_location_incompatible(o) for o in opps):
        return 1, ["fall/co-op internship outside LA/remote"]
    if any(_is_internship_role(o) and not _is_summer_internship(o) for o in opps):
        return 6, ["internship discovered"]
    if any(_is_summer_internship(o) for o in opps):
        return 3, ["summer internship discovered"]
    return 5, ["roles discovered"]


def _score_relationship_depth(contacts: list[ContactRecord]) -> tuple[int, str]:
    """Daily momentum score. Active conversation should jump the action queue."""
    replied = sum(1 for c in contacts if c.status in REPLIED_STATUSES)
    accepted = sum(1 for c in contacts if c.status in ACCEPTED_STATUSES)

    if replied >= 3:
        return 20, f"{replied} warm contacts"
    if replied >= 1:
        return 15, f"{replied} warm contact(s)"
    if accepted >= 3:
        return 12, f"{accepted} accepted"
    if accepted >= 1:
        return 8, f"{accepted} accepted"
    return 0, ""


def _score_relationship_priority(contacts: list[ContactRecord]) -> tuple[int, str]:
    """Small strategic score bonus for traction without letting traction dominate tiering."""
    replied = sum(1 for c in contacts if c.status in REPLIED_STATUSES)
    accepted = sum(1 for c in contacts if c.status in ACCEPTED_STATUSES)

    if replied >= 1:
        return 6, f"{replied} warm contact(s)"
    if accepted >= 3:
        return 4, f"{accepted} accepted"
    if accepted >= 1:
        return 3, f"{accepted} accepted"
    return 0, ""


def _derive_stage(
    account_score: int,
    contacts: list[ContactRecord],
) -> tuple[str, str, str]:
    if not contacts:
        if account_score >= 30:
            return "priority_target", "Map contacts on LinkedIn", \
                (date.today() + timedelta(days=7)).isoformat()
        return "unqualified", "Evaluate fit", \
            (date.today() + timedelta(days=14)).isoformat()

    invites = sum(1 for c in contacts if c.status in INVITED_STATUSES)
    accepted = sum(1 for c in contacts if c.status in ACCEPTED_STATUSES)
    replied = sum(1 for c in contacts if c.status in REPLIED_STATUSES)

    if replied > 0:
        return "conversation_started", "Continue conversation; push for coffee chat", \
            (date.today() + timedelta(days=3)).isoformat()
    if accepted > 0:
        return "connected_no_conversation", "Draft and send follow-up", \
            (date.today() + timedelta(days=2)).isoformat()
    if invites > 0:
        return "outreach_active", "Reconcile LinkedIn; await accepts", \
            (date.today() + timedelta(days=5)).isoformat()
    return "people_mapped", "Send LinkedIn invites", \
        (date.today() + timedelta(days=3)).isoformat()


def _campaign_plan_for_account(row: AccountRow) -> tuple[str, str, int, str, str]:
    flags = {item.strip() for item in row.data_quality_flags.split(";") if item.strip()}
    if "needs_domain_enrichment" in flags or "role_without_domain_context" in flags:
        if row.fit_score < 25 and row.score_brand == 0 and row.score_role_fit < 18:
            return (
                "pause_account",
                "none",
                10,
                "Company/domain fit is under-specified and there is not enough role, brand, or relationship signal to enrich yet.",
                "lane_1_allowed",
            )
        return (
            "enrich_company_context",
            "research",
            85 if row.tier == "A" else (65 if row.tier == "B" else 45),
            "Company/domain fit is under-specified for a relationship campaign; enrich before more touches.",
            "fresh_role_only",
        )
    if row.tier in {"C", "L3"}:
        return (
            "pause_account",
            "none",
            10,
            "Not in Relationship A/B or Large L1/L2 priority; do not spend account-campaign budget right now.",
            "lane_1_allowed",
        )
    if row.account_score < 18:
        return (
            "pause_account",
            "none",
            10,
            "Not enough durable account fit to spend relationship-engine budget.",
            "lane_1_allowed",
        )
    if row.account_stage == "conversation_started":
        return (
            "continue_conversation",
            "linkedin",
            100,
            "A real conversation exists; prioritize moving toward coffee chat, referral, or routing.",
            "track_2_owns",
        )
    if row.account_stage == "connected_no_conversation":
        return (
            "follow_up_connected_contact",
            "linkedin",
            95,
            "Someone accepted but has not replied; send the accepted-invite follow-up before new outreach.",
            "track_2_owns",
        )
    if row.account_stage == "people_mapped" and row.invites_sent == 0:
        if row.email_contacts > 0 and row.tier in {"A", "L1"}:
            return (
                "send_initial_multichannel_outreach",
                "linkedin+email",
                86 if row.tier == "A" else 76,
                "Relevant people are mapped and at least one email is available; run LinkedIn and email in parallel.",
                "track_2_owns",
            )
        return (
            "send_initial_invites",
            "linkedin",
            80 if row.tier == "A" else 60,
            "Relevant people are mapped but no LinkedIn wave has been sent.",
            "track_2_owns",
        )
    if row.people_mapped < MIN_MAPPED_CONTACTS:
        return (
            "map_more_contacts",
            "linkedin+email_research" if row.tier in {"A", "L1"} else "linkedin",
            75 if row.tier == "A" else 55,
            f"Only {row.people_mapped} relevant contact(s) mapped; build a better account map before sending more.",
            "track_2_owns",
        )
    if row.account_stage == "outreach_active":
        if row.invites_sent >= LINKEDIN_WAVE_SIZE and row.accepted == 0 and row.replies == 0:
            if row.email_contacts > 0:
                return (
                    "send_cold_email_followup",
                    "email",
                    92 if row.tier == "A" else 72,
                    f"{row.invites_sent} LinkedIn invites with no accepts/replies and email exists; switch the next touch to email.",
                    "track_2_owns",
                )
            return (
                "find_email_path",
                "email_research",
                90 if row.tier == "A" else 70,
                f"{row.invites_sent} LinkedIn invites with no accepts/replies; find an email/contact-info path instead of another blind wave.",
                "track_2_owns",
            )
        if row.invites_sent < LINKEDIN_WAVE_SIZE:
            return (
                "expand_linkedin_wave",
                "linkedin",
                78 if row.tier == "A" else 58,
                f"{row.invites_sent} invite(s) sent; expand toward an {LINKEDIN_WAVE_SIZE}-person wave if contacts are relevant.",
                "track_2_owns",
            )
        return (
            "wait_for_accepts",
            "linkedin",
            45,
            "LinkedIn wave is in flight; reconcile accepts/replies before new account touches.",
            "track_2_owns",
        )
    if row.account_stage in {"priority_target", "unqualified"}:
        return (
            "map_more_contacts" if row.account_score >= 30 else "enrich_company_context",
            (
                "linkedin+email_research"
                if row.account_score >= 30 and row.tier in {"A", "L1"}
                else ("linkedin" if row.account_score >= 30 else "research")
            ),
            70 if row.tier == "A" else 50,
            "Account has enough score to inspect, but is not ready for a campaign touch yet.",
            "fresh_role_only" if row.tier in {"A", "B"} else "lane_1_allowed",
        )
    return (
        "pause_account",
        "none",
        10,
        "No high-leverage campaign action detected.",
        "lane_1_allowed",
    )


def _daily_action_priority_for_account(row: AccountRow) -> int:
    """Execution urgency for today's queue, separate from durable account rank."""
    action_floor = row.campaign_priority
    tier_bonus = 5 if row.tier == "A" else (2 if row.tier == "B" else 0)
    score_bonus = min(5, row.account_score // 12)
    momentum_bonus = min(6, row.score_relationship_momentum // 3)
    return max(row.campaign_priority, action_floor + tier_bonus + score_bonus + momentum_bonus)


def _account_rank_key(row: AccountRow) -> tuple[int, int, str]:
    return (row.account_score, row.fit_score, row.company.lower())


def _is_tier_a_eligible(row: AccountRow) -> bool:
    return (
        row.account_score >= TIER_A_MIN_SCORE
        and row.account_track in TIER_A_TRACK_QUOTAS
    )


def _assign_segmented_tiers(rows: list[AccountRow]) -> None:
    """Assign relationship tiers and separate large-company priority labels."""
    rows.sort(key=_account_rank_key, reverse=True)
    for row in rows:
        row.tier = "C"

    selected: set[str] = set()
    for track, quota in TIER_A_TRACK_QUOTAS.items():
        candidates = [
            row
            for row in rows
            if row.account_track == track and _is_tier_a_eligible(row)
        ]
        for row in candidates[:quota]:
            row.tier = "A"
            selected.add(row.organization_id)

    remaining_slots = RELATIONSHIP_TIER_A_TOTAL - len(selected)
    if remaining_slots > 0:
        fillers = [
            row
            for row in rows
            if row.organization_id not in selected and _is_tier_a_eligible(row)
        ]
        for row in fillers:
            row.tier = "A"
            selected.add(row.organization_id)
            remaining_slots -= 1
            if remaining_slots <= 0:
                break

    large_companies = [row for row in rows if row.account_track == "Large Company"]
    for row in large_companies[:LARGE_COMPANY_L1_TOTAL]:
        row.tier = "L1"
    for row in large_companies[LARGE_COMPANY_L1_TOTAL:LARGE_COMPANY_L1_TOTAL + LARGE_COMPANY_L2_TOTAL]:
        row.tier = "L2"
    for row in large_companies[LARGE_COMPANY_L1_TOTAL + LARGE_COMPANY_L2_TOTAL:]:
        row.tier = "L3"

    tier_b_assigned = 0
    for row in rows:
        if row.organization_id in selected:
            continue
        if row.account_track == "Large Company":
            continue
        if row.account_score < TIER_B_MIN_SCORE:
            continue
        row.tier = "B"
        tier_b_assigned += 1
        if tier_b_assigned >= TIER_B_TOTAL:
            break


def _target_tags(row: AccountRow) -> set[str]:
    return _split_normalized_tags(" ".join([row.target_lists, row.why_fit, row.tags]))


def _is_strategic_wishlist(row: AccountRow) -> bool:
    tags = _target_tags(row)
    if tags.intersection(MANUAL_PRIORITY_TAGS):
        return True
    text = " ".join([row.target_lists, row.why_fit]).lower()
    return any(
        _mentions_domain_term(text, token)
        for token in ("manual priority", "tier a", "dream company", "priority account")
    )


def _needs_context_enrichment(row: AccountRow) -> bool:
    flags = {item.strip() for item in row.data_quality_flags.split(";") if item.strip()}
    return bool(
        flags.intersection(
            {
                "needs_domain_enrichment",
                "role_without_domain_context",
                "context_inferred_from_job",
                "team_size_unparsed",
            }
        )
    )


def _account_track_for_row(row: AccountRow) -> str:
    """Operational account segment. Strategic wishlist and enrichment are workbook views."""
    if row.team_size is not None and row.team_size >= 1000:
        return "Large Company"
    if row.team_size is not None and row.team_size >= 200:
        return "Growth / Mid-Market"
    tags = _target_tags(row)
    if row.org_type.lower() == "startup" or tags.intersection({"yc", "startup", "hiring"}):
        return "Startup / Founder-Led"
    if row.team_size is not None and row.team_size < 200:
        return "Startup / Founder-Led"
    if _needs_context_enrichment(row):
        return "Needs Enrichment"
    return "Growth / Mid-Market"


# Human-readable label map for why_fit
_LABEL = {
    "artificial-intelligence": "AI/ML",
    "ai": "AI/ML",
    "machine-learning": "AI/ML",
    "generative-ai": "gen AI",
    "applied-ai": "applied AI",
    "llm": "LLM/AI",
    "agent": "agent AI",
    "conversational-ai": "conversational AI",
    "nlp": "NLP",
    "data": "data platform",
    "data-infrastructure": "data infra",
    "data-platform": "data platform",
    "data-pipeline": "data pipelines",
    "etl": "ETL/data",
    "analytics": "analytics",
    "business-intelligence": "BI/analytics",
    "warehousing": "data warehouse",
    "integration": "API/integration",
    "api": "API/integration",
    "connectivity": "API/connectivity",
    "observability": "observability",
    "monitoring": "monitoring",
    "developer-tools": "dev tools",
    "developer-experience": "DevEx",
    "developer-platform": "dev platform",
    "infrastructure": "infra",
    "hiring": "hiring tech",
    "recruiting": "recruiting tech",
    "hr-tech": "HR tech",
    "talent": "talent",
    "workflow-automation": "workflow automation",
    "automation": "automation",
    "marketplace": "marketplace",
    "logistics": "logistics",
    "mobility": "mobility",
    "transportation": "transportation",
    "delivery": "delivery",
    "gig-economy": "gig economy",
    "fintech": "fintech",
    "payments": "payments",
    "financial-technology": "fintech",
    "billing": "billing/fintech",
    "banking": "banking/fintech",
    "healthcare": "healthcare",
    "health-tech": "health tech",
    "healthtech": "health tech",
    "digital-health": "digital health",
    "medtech": "medtech",
    "agentic": "AI agents",
    "copilot": "AI copilot",
    "productivity": "productivity",
    "saas": "SaaS",
}


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_account_rows(workbook_dir: Path) -> list[AccountRow]:
    wb = OutreachWorkbook(workbook_dir)
    orgs = wb.list_organizations()
    contacts = wb.list_contacts()
    opps = wb.list_opportunities()

    contacts_by_org: dict[str, list[ContactRecord]] = defaultdict(list)
    for c in contacts:
        contacts_by_org[c.organization_id].append(c)

    opps_by_org: dict[str, list[OpportunityRecord]] = defaultdict(list)
    for o in opps:
        opps_by_org[o.organization_id].append(o)

    rows: list[AccountRow] = []

    for org in orgs:
        oc = contacts_by_org[org.organization_id]
        company_contacts = _company_relevant_contacts(org, oc)
        oo = opps_by_org[org.organization_id]

        s_profile, p_signals, tags_str, team_size = _score_profile_fit(org)
        s_role, best_role, hiring_signal = _score_role_fit(oo)
        s_team = _score_team_gate(team_size)
        s_reach, r_signals = _score_reachability(org, company_contacts)
        s_hiring, _ = _score_hiring(oo)
        s_rel_momentum, momentum_label = _score_relationship_depth(company_contacts)
        s_rel_priority, rel_label = _score_relationship_priority(company_contacts)
        quality_flags = _data_quality_flags(
            org=org,
            profile_fit=s_profile,
            team_size=team_size,
            opportunities=oo,
        )
        s_brand, brand_label = _score_brand(org)
        s_pitch, pitch_label = _score_pitch_strength(
            org=org,
            profile_fit=s_profile,
            reachability=s_reach,
            team_size=team_size,
            data_quality_flags=quality_flags,
        )
        s_account_hiring = _score_account_hiring(s_role, s_hiring)

        # No-domain-data discount: jobs.xlsx imports with empty tags/description
        # get role+hiring credit but we have no idea if they're actually a fit.
        # TODO(data): fix root cause by importing company tags/description from
        # ResumeGenerator v1 jobs.xlsx at import time, or pull from YC/BuiltIn
        # discovery for the same company. Until then, -8 penalty applies.
        no_domain_data = (s_profile == 0 and "imported from resumegenerator" in org.notes.lower())
        s_domain_penalty = -8 if no_domain_data else 0

        fit_score = max(0, s_profile + s_role + s_team + s_reach + s_hiring
                        + s_rel_momentum + s_domain_penalty)
        account_score = _score_account_campaign(
            profile_fit=s_profile,
            reachability=s_reach,
            brand=s_brand,
            relationship=s_rel_priority,
            account_hiring=s_account_hiring,
            pitch_strength=s_pitch,
            team_gate=s_team,
            data_quality_flags=quality_flags,
        )

        # Build why_fit
        seen: set[str] = set()
        labels: list[str] = []
        for s in p_signals:
            lbl = _LABEL.get(s, s)
            if lbl not in seen:
                seen.add(lbl)
                labels.append(lbl)
        labels.extend(r_signals)
        if brand_label:
            labels.append(brand_label)
        if pitch_label:
            labels.append(pitch_label)
        if rel_label:
            labels.insert(0, rel_label)
        elif momentum_label:
            labels.insert(0, momentum_label)
        if s_hiring > 0:
            labels.append(hiring_signal)
        why_fit = "; ".join(list(dict.fromkeys(labels))[:6])

        invites_sent = sum(1 for c in company_contacts if c.status in INVITED_STATUSES)
        email_contacts = sum(1 for c in company_contacts if c.email.strip())
        accepted = sum(1 for c in company_contacts if c.status in ACCEPTED_STATUSES)
        replies = sum(1 for c in company_contacts if c.status in REPLIED_STATUSES)

        stage, next_action, next_due = _derive_stage(account_score, company_contacts)
        row = AccountRow(
            organization_id=org.organization_id,
            company=org.name,
            org_type=org.organization_type,
            city=org.city,
            website=org.website,
            fit_score=fit_score,
            account_score=account_score,
            why_fit=why_fit,
            target_role=best_role or "—",
            hiring_signal=hiring_signal,
            people_mapped=len(company_contacts),
            email_contacts=email_contacts,
            invites_sent=invites_sent,
            accepted=accepted,
            replies=replies,
            account_stage=stage,
            next_action=next_action,
            next_due_date=next_due,
            target_lists=org.target_lists,
            team_size=team_size,
            tags=tags_str,
            score_profile_fit=s_profile,
            score_role_fit=s_role,
            score_team_gate=s_team,
            score_reachability=s_reach,
            score_hiring=s_hiring,
            score_relationship=s_rel_priority,
            score_relationship_momentum=s_rel_momentum,
            score_brand=s_brand,
            score_pitch_strength=s_pitch,
            score_account_hiring=s_account_hiring,
            data_quality_flags=";".join(quality_flags),
        )
        (
            row.campaign_action,
            row.campaign_channel,
            row.campaign_priority,
            row.campaign_reason,
            row.lane_1_policy,
        ) = _campaign_plan_for_account(row)
        row.daily_action_priority = _daily_action_priority_for_account(row)
        row.account_track = _account_track_for_row(row)
        rows.append(row)

    _assign_segmented_tiers(rows)
    for r in rows:
        (
            r.campaign_action,
            r.campaign_channel,
            r.campaign_priority,
            r.campaign_reason,
            r.lane_1_policy,
        ) = _campaign_plan_for_account(r)
        r.daily_action_priority = _daily_action_priority_for_account(r)
        r.account_track = _account_track_for_row(r)

    return rows


def build_campaign_plan_rows(rows: list[AccountRow]) -> list[CampaignPlanRow]:
    actionable = [
        row
        for row in rows
        if row.campaign_action not in {"pause_account", "wait_for_accepts"}
    ]
    actionable.sort(
        key=lambda row: (
            row.daily_action_priority,
            row.tier == "A",
            row.account_score,
            row.fit_score,
            row.company.lower(),
        ),
        reverse=True,
    )
    return [
        CampaignPlanRow(
            company=row.company,
            tier=row.tier,
            fit_score=row.fit_score,
            account_score=row.account_score,
            account_stage=row.account_stage,
            campaign_action=row.campaign_action,
            campaign_channel=row.campaign_channel,
            account_track=row.account_track,
            campaign_priority=row.campaign_priority,
            daily_action_priority=row.daily_action_priority,
            campaign_reason=row.campaign_reason,
            lane_1_policy=row.lane_1_policy,
            why_fit=row.why_fit,
            people_mapped=row.people_mapped,
            email_contacts=row.email_contacts,
            invites_sent=row.invites_sent,
            accepted=row.accepted,
            replies=row.replies,
            target_role=row.target_role,
            next_due_date=row.next_due_date,
            organization_id=row.organization_id,
        )
        for row in actionable
    ]


def build_track_2_daily_plan(
    rows: list[AccountRow],
    *,
    budget: DailyPlanBudget = DailyPlanBudget(),
) -> dict[str, object]:
    campaign_rows = build_campaign_plan_rows(rows)
    selected: list[DailyPlanItem] = []
    skipped: list[DailyPlanItem] = []
    used = {
        "total_actions": 0,
        "companies": 0,
        "linkedin_invites": 0,
        "linkedin_followups": 0,
        "company_mapping": 0,
        "email_research": 0,
        "context_enrichment": 0,
        "email_drafts": 0,
    }
    seen_companies: set[str] = set()

    for row in campaign_rows:
        item = _daily_plan_item_for_campaign(row)
        skip_reason = _daily_plan_skip_reason(item, used, seen_companies, budget)
        if skip_reason:
            item.skip_reason = skip_reason
            skipped.append(item)
            continue
        selected.append(item)
        seen_companies.add(item.organization_id or item.company.lower())
        used["total_actions"] += 1
        used["companies"] += 1
        used["linkedin_invites"] += item.expected_linkedin_invites
        used["linkedin_followups"] += item.expected_linkedin_followups
        used["company_mapping"] += item.expected_company_mapping
        used["email_research"] += item.expected_email_research
        used["context_enrichment"] += item.expected_context_enrichment
        used["email_drafts"] += item.expected_email_drafts

    selected.sort(key=lambda item: (item.phase_order, -item.daily_action_priority, item.company.lower()))
    skipped.sort(key=lambda item: (item.phase_order, -item.daily_action_priority, item.company.lower()))
    return {
        "budget": budget.__dict__,
        "used": used,
        "selected_count": len(selected),
        "skipped_count": len(skipped),
        "selected": [item.__dict__ for item in selected],
        "skipped": [item.__dict__ for item in skipped],
        "summary": _daily_plan_summary(selected),
        "phase_summary": _daily_plan_phase_summary(selected),
        "execution_order": _daily_plan_execution_order(),
    }


def _daily_plan_item_for_campaign(row: CampaignPlanRow) -> DailyPlanItem:
    item = DailyPlanItem(
        company=row.company,
        tier=row.tier,
        campaign_action=row.campaign_action,
        campaign_channel=row.campaign_channel,
        daily_action_priority=row.daily_action_priority,
        account_score=row.account_score,
        reason=row.campaign_reason,
        organization_id=row.organization_id,
    )
    remaining_wave = max(0, LINKEDIN_WAVE_SIZE - row.invites_sent)
    if row.campaign_action in {"continue_conversation", "follow_up_connected_contact"}:
        item.expected_linkedin_followups = 1
    elif row.campaign_action == "send_initial_invites":
        item.expected_linkedin_invites = min(remaining_wave or 3, 3)
    elif row.campaign_action == "expand_linkedin_wave":
        item.expected_linkedin_invites = min(remaining_wave or 1, 3)
    elif row.campaign_action == "send_initial_multichannel_outreach":
        item.expected_linkedin_invites = 1
        item.expected_email_drafts = 1
    elif row.campaign_action == "send_cold_email_followup":
        item.expected_email_drafts = 1
    elif row.campaign_action == "map_more_contacts":
        item.expected_company_mapping = 1
        if "email" in row.campaign_channel:
            item.expected_email_research = 1
    elif row.campaign_action == "find_email_path":
        item.expected_email_research = 1
    elif row.campaign_action == "enrich_company_context":
        item.expected_context_enrichment = 1
    item.phase, item.phase_order, item.can_parallelize = _daily_plan_phase(item)
    return item


def _daily_plan_phase(item: DailyPlanItem) -> tuple[str, int, bool]:
    action = item.campaign_action
    if action == "continue_conversation":
        return "1_continue_live_conversations", 10, False
    if action == "follow_up_connected_contact":
        return "2_follow_up_warm_accepts", 20, False
    if action == "find_email_path" and item.expected_email_research:
        return "3_contact_and_email_research", 30, True
    if action == "map_more_contacts":
        return "4_contact_mapping", 40, True
    if action in {"send_initial_invites", "expand_linkedin_wave", "send_initial_multichannel_outreach"}:
        return "5_send_linkedin_invites", 50, False
    if action == "send_cold_email_followup":
        return "6_draft_email_touch", 60, False
    if action == "enrich_company_context":
        return "7_context_enrichment", 70, True
    return "9_other", 90, False


def _daily_plan_skip_reason(
    item: DailyPlanItem,
    used: dict[str, int],
    seen_companies: set[str],
    budget: DailyPlanBudget,
) -> str:
    company_key = item.organization_id or item.company.lower()
    if company_key in seen_companies:
        return "company_already_selected"
    if used["total_actions"] + 1 > budget.max_total_actions:
        return "daily_action_budget_exhausted"
    if used["companies"] + 1 > budget.max_companies:
        return "company_budget_exhausted"
    checks = [
        ("linkedin_invites", item.expected_linkedin_invites, budget.max_linkedin_invites),
        ("linkedin_followups", item.expected_linkedin_followups, budget.max_linkedin_followups),
        ("company_mapping", item.expected_company_mapping, budget.max_company_mapping),
        ("email_research", item.expected_email_research, budget.max_email_research),
        ("context_enrichment", item.expected_context_enrichment, budget.max_context_enrichment),
        ("email_drafts", item.expected_email_drafts, budget.max_email_drafts),
    ]
    for key, increment, cap in checks:
        if increment and used[key] + increment > cap:
            return f"{key}_budget_exhausted"
    return ""


def _daily_plan_summary(items: list[DailyPlanItem]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for item in items:
        summary[item.campaign_action] = summary.get(item.campaign_action, 0) + 1
    return summary


def _daily_plan_phase_summary(items: list[DailyPlanItem]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for item in items:
        summary[item.phase] = summary.get(item.phase, 0) + 1
    return summary


def _daily_plan_execution_order() -> list[dict[str, object]]:
    return [
        {
            "phase": "1_continue_live_conversations",
            "order": 10,
            "parallelizable": False,
            "notes": "Handle active replies first; do not let fresh outreach bury live conversations.",
        },
        {
            "phase": "2_follow_up_warm_accepts",
            "order": 20,
            "parallelizable": False,
            "notes": "Send accepted-invite follow-ups before starting new waves.",
        },
        {
            "phase": "3_contact_and_email_research",
            "order": 30,
            "parallelizable": True,
            "notes": "Inspect LinkedIn Contact Info and company/contact paths; store emails before drafting email.",
        },
        {
            "phase": "4_contact_mapping",
            "order": 40,
            "parallelizable": True,
            "notes": "Map more relevant people for priority accounts that do not yet have a good account map.",
        },
        {
            "phase": "5_send_linkedin_invites",
            "order": 50,
            "parallelizable": False,
            "notes": "Send bounded LinkedIn invite waves only after warm replies/follow-ups and required research tasks.",
        },
        {
            "phase": "6_draft_email_touch",
            "order": 60,
            "parallelizable": False,
            "notes": "Draft email touches only when the comm engine is ready and daily email budget is nonzero.",
        },
        {
            "phase": "7_context_enrichment",
            "order": 70,
            "parallelizable": True,
            "notes": "Enrich company context in the background; it should not block live relationship work.",
        },
    ]


def audit_track_2_core(rows: list[AccountRow]) -> dict[str, object]:
    action_counts: dict[str, int] = {}
    channel_counts: dict[str, int] = {}
    issue_counts: dict[str, int] = {}
    issues: list[dict[str, object]] = []
    priority_tiers = {"A", "B", "L1", "L2"}
    actionable_actions = {
        "continue_conversation",
        "follow_up_connected_contact",
        "send_initial_multichannel_outreach",
        "send_initial_invites",
        "map_more_contacts",
        "send_cold_email_followup",
        "find_email_path",
        "expand_linkedin_wave",
        "enrich_company_context",
    }

    def flag(row: AccountRow, code: str, detail: str) -> None:
        issue_counts[code] = issue_counts.get(code, 0) + 1
        issues.append(
            {
                "company": row.company,
                "tier": row.tier,
                "account_score": row.account_score,
                "campaign_action": row.campaign_action,
                "campaign_channel": row.campaign_channel,
                "code": code,
                "detail": detail,
            }
        )

    for row in rows:
        action_counts[row.campaign_action] = action_counts.get(row.campaign_action, 0) + 1
        channel_counts[row.campaign_channel] = channel_counts.get(row.campaign_channel, 0) + 1
        if row.tier in priority_tiers and not row.campaign_action:
            flag(row, "missing_action", "Priority account has no campaign action.")
        if row.campaign_action in actionable_actions and not row.campaign_channel:
            flag(row, "missing_channel", "Actionable account has no channel.")
        if row.campaign_action == "send_initial_multichannel_outreach" and row.email_contacts <= 0:
            flag(row, "parallel_without_email", "Parallel outreach requires at least one email contact.")
        if row.campaign_action == "send_cold_email_followup" and row.email_contacts <= 0:
            flag(row, "email_without_email", "Cold email follow-up requires at least one email contact.")
        if row.campaign_action == "map_more_contacts" and row.people_mapped >= MIN_MAPPED_CONTACTS:
            flag(row, "mapping_when_enough_contacts", "Account already has enough mapped contacts.")
        if row.campaign_action == "pause_account" and row.tier in {"A", "B", "L1", "L2"}:
            flag(row, "priority_paused", "Priority account is paused; verify this is intentional.")

    return {
        "total_accounts": len(rows),
        "priority_accounts": sum(1 for row in rows if row.tier in priority_tiers),
        "action_counts": action_counts,
        "channel_counts": channel_counts,
        "issue_counts": issue_counts,
        "issues": issues,
        "is_clean": not issues,
    }


# ---------------------------------------------------------------------------
# Excel generator
# ---------------------------------------------------------------------------

COLUMNS: list[tuple[str, str, int]] = [
    ("Company",         "company",          22),
    ("Tier",            "tier",              6),
    ("Account Score",   "account_score",    13),
    ("Fit Score",       "fit_score",         10),
    ("Why Fit",         "why_fit",           45),
    ("Account Stage",   "account_stage",     24),
    ("Target Role",     "target_role",       28),
    ("Hiring Signal",   "hiring_signal",     22),
    ("Team Size",       "team_size",         10),
    ("City",            "city",              18),
    ("People Mapped",   "people_mapped",     14),
    ("Email Contacts",  "email_contacts",    14),
    ("Invites Sent",    "invites_sent",      13),
    ("Accepted",        "accepted",          10),
    ("Replies",         "replies",            9),
    ("Coffee Chats",    "coffee_chats",      13),
    ("Advocates",       "advocates",         11),
    ("Next Action",     "next_action",       40),
    ("Next Due",        "next_due_date",     12),
    ("Campaign Action", "campaign_action",   26),
    ("Campaign Channel","campaign_channel",  18),
    ("Campaign Priority","campaign_priority",16),
    ("Daily Action Priority","daily_action_priority",20),
    ("Campaign Reason", "campaign_reason",   50),
    ("Lane 1 Policy",   "lane_1_policy",     18),
    ("Account Track",   "account_track",     22),
    ("Score: Profile",  "score_profile_fit", 14),
    ("Score: Role",     "score_role_fit",    12),
    ("Score: Team",     "score_team_gate",   12),
    ("Score: Reach",    "score_reachability",13),
    ("Score: Hiring",   "score_hiring",      13),
    ("Score: Rel",      "score_relationship",11),
    ("Score: Momentum", "score_relationship_momentum",16),
    ("Score: Brand",    "score_brand",       13),
    ("Score: Pitch",    "score_pitch_strength",13),
    ("Score: Account Hiring","score_account_hiring",18),
    ("Org Type",        "org_type",          12),
    ("Tags",            "tags",              40),
    ("Data Flags",      "data_quality_flags",30),
    ("Target Lists",    "target_lists",      20),
    ("Website",         "website",           30),
]

ACTION_COLS: list[tuple[str, str, int]] = [
    ("Company",       "company",       22),
    ("Tier",          "tier",           6),
    ("Account Score", "account_score", 13),
    ("Fit Score",     "fit_score",     10),
    ("Why Fit",       "why_fit",       45),
    ("Account Stage", "account_stage", 24),
    ("Accepted",      "accepted",      10),
    ("Replies",       "replies",        9),
    ("Next Action",   "next_action",   40),
    ("Next Due",      "next_due_date", 12),
    ("People Mapped", "people_mapped", 14),
    ("Email Contacts","email_contacts", 14),
    ("Invites Sent",  "invites_sent",  13),
]

CAMPAIGN_COLS: list[tuple[str, str, int]] = [
    ("Company",           "company",            22),
    ("Tier",              "tier",                6),
    ("Daily Priority",    "daily_action_priority", 14),
    ("Base Priority",     "campaign_priority",  13),
    ("Campaign Action",   "campaign_action",    26),
    ("Channel",           "campaign_channel",   16),
    ("Account Track",     "account_track",      22),
    ("Account Stage",     "account_stage",      24),
    ("Lane 1 Policy",     "lane_1_policy",      18),
    ("Reason",            "campaign_reason",    54),
    ("Why Fit",           "why_fit",            45),
    ("Account Score",     "account_score",      13),
    ("Fit Score",         "fit_score",          10),
    ("Target Role",       "target_role",        28),
    ("People Mapped",     "people_mapped",      14),
    ("Email Contacts",    "email_contacts",     14),
    ("Invites Sent",      "invites_sent",       13),
    ("Accepted",          "accepted",           10),
    ("Replies",           "replies",             9),
    ("Next Due",          "next_due_date",      12),
    ("Organization ID",    "organization_id",    24),
]

TIER_COLORS = {
    "A": "D6F5D6",
    "B": "FFF9C4",
    "C": "F5F5F5",
    "L1": "D9EAF7",
    "L2": "EAF3F8",
    "L3": "F5F5F5",
}
STAGE_COLORS = {
    "conversation_started":      "C8E6C9",
    "connected_no_conversation": "B3E0F7",
    "outreach_active":           "E3F2FD",
    "people_mapped":             "FFF9C4",
    "priority_target":           "FFE0CC",
    "unqualified":               "F5F5F5",
    "coffee_chat":               "A5D6A7",
    "warm_advocate":             "81C784",
    "referral_path":             "66BB6A",
    "paused":                    "EEEEEE",
}
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
CTR  = Alignment(horizontal="center", vertical="top")
LEFT = Alignment(horizontal="left",   vertical="top")
WRAP = Alignment(horizontal="left",   vertical="top", wrap_text=True)


def _write_header(ws, cols: list) -> None:
    for ci, (hdr, _, w) in enumerate(cols, 1):
        c = ws.cell(1, ci, hdr)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CTR
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20


def _write_row(ws, ri: int, row: AccountRow, cols: list) -> None:
    tier_fill  = PatternFill("solid", fgColor=TIER_COLORS.get(row.tier, "F5F5F5"))
    stage_fill = PatternFill("solid", fgColor=STAGE_COLORS.get(row.account_stage, "F5F5F5"))
    for ci, (hdr, attr, _) in enumerate(cols, 1):
        val = getattr(row, attr, "")
        if val is None:
            val = ""
        c = ws.cell(ri, ci, val)
        c.font = BODY_FONT
        if hdr == "Tier":
            c.fill = tier_fill
            c.alignment = CTR
        elif hdr == "Account Stage":
            c.fill = stage_fill
            c.alignment = LEFT
        elif hdr in ("Why Fit", "Next Action", "Tags", "Campaign Reason", "Reason", "Data Flags"):
            c.alignment = WRAP
        elif hdr in ("Account Score", "Fit Score", "People Mapped", "Email Contacts", "Invites Sent", "Accepted",
                     "Replies", "Coffee Chats", "Advocates", "Team Size",
                     "Score: Profile", "Score: Role", "Score: Team",
                     "Score: Reach", "Score: Hiring", "Score: Rel", "Score: Momentum", "Score: Brand",
                     "Score: Pitch", "Score: Account Hiring",
                     "Campaign Priority", "Daily Action Priority", "Daily Priority",
                     "Base Priority", "Priority"):
            c.alignment = CTR
        else:
            c.alignment = LEFT
    ws.row_dimensions[ri].height = 16


def generate_excel(rows: list[AccountRow], output_path: Path) -> Path:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Account Tracker"
    ws1.freeze_panes = "C2"
    _write_header(ws1, COLUMNS)
    for ri, row in enumerate(rows, 2):
        _write_row(ws1, ri, row, COLUMNS)
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    ws2 = wb.create_sheet("Tier A — Active Campaign")
    ws2.freeze_panes = "C2"
    _write_header(ws2, COLUMNS)
    for ri, row in enumerate((r for r in rows if r.tier == "A"), 2):
        _write_row(ws2, ri, row, COLUMNS)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    ws3 = wb.create_sheet("Action Queue")
    ws3.freeze_panes = "B2"
    _write_header(ws3, ACTION_COLS)
    ACTIONABLE = {
        "conversation_started", "connected_no_conversation",
        "outreach_active", "people_mapped", "priority_target",
    }
    STAGE_PRI = {
        "conversation_started": 0,
        "connected_no_conversation": 1,
        "outreach_active": 2,
        "people_mapped": 3,
        "priority_target": 4,
    }
    action_rows = sorted(
        [r for r in rows if r.tier in ("A", "B", "L1", "L2") and r.account_stage in ACTIONABLE],
        key=lambda r: (STAGE_PRI.get(r.account_stage, 9), r.tier, -r.account_score, -r.fit_score),
    )
    for ri, row in enumerate(action_rows, 2):
        _write_row(ws3, ri, row, ACTION_COLS)
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(ACTION_COLS))}1"

    ws4 = wb.create_sheet("Campaign Plan")
    ws4.freeze_panes = "D2"
    _write_header(ws4, CAMPAIGN_COLS)
    for ri, row in enumerate(build_campaign_plan_rows(rows), 2):
        _write_row(ws4, ri, row, CAMPAIGN_COLS)
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(CAMPAIGN_COLS))}1"

    view_specs = [
        ("Startup Founder-Led", lambda r: r.account_track == "Startup / Founder-Led"),
        ("Growth Mid-Market", lambda r: r.account_track == "Growth / Mid-Market"),
        ("Large Company", lambda r: r.account_track == "Large Company"),
        ("Large Company Priority", lambda r: r.tier in {"L1", "L2"}),
        ("Strategic Wishlist", _is_strategic_wishlist),
        ("Needs Enrichment", _needs_context_enrichment),
    ]
    for sheet_name, predicate in view_specs:
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "C2"
        _write_header(ws, COLUMNS)
        view_rows = [row for row in rows if predicate(row)]
        for ri, row in enumerate(view_rows, 2):
            _write_row(ws, ri, row, COLUMNS)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run(workbook_dir: Path, output_path: Path) -> tuple[list[AccountRow], Path]:
    rows = build_account_rows(workbook_dir)
    path = generate_excel(rows, output_path)
    return rows, path
