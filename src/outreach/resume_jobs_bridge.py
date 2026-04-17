from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable

from outreach.tracking import OpportunityType, OrganizationType, SourceKind

DEFAULT_INCLUDE_STATUSES = ("queued", "generated", "applied")
DEFAULT_TARGET_LISTS = "jobs;resume_generator;pre_apply"
DEFAULT_COMPANY_OVERRIDES_FILENAME = "company_overrides.csv"

STARTUP_BIAS_SCORES = {
    "high": 2.0,
    "medium": 1.0,
    "low": 0.5,
    "neutral": 0.0,
    "none": 0.0,
    "deprioritize": -1.0,
}

COMPANY_TYPE_SCORES = {
    "startup": 1.5,
    "growth": 1.0,
    "big_company": -0.5,
    "unknown": 0.0,
    "company": 0.0,
}


@dataclass(frozen=True)
class ResumeJob:
    row_id: str
    company: str
    role_title: str
    location: str
    url: str
    url_hash: str
    source: str
    status: str
    normalized_status: str
    fit_score: float | None
    fit_rationale: str
    date_found: date | None
    date_posted_raw: str
    folder_path: str
    jd_text: str
    notes: str


@dataclass(frozen=True)
class ResumeImportSelection:
    jobs: list[ResumeJob]
    skipped_missing_identity: int
    skipped_status: int
    skipped_score: int
    skipped_age: int
    duplicates_removed: int


@dataclass(frozen=True)
class CompanyOverride:
    company: str
    normalized_company: str
    company_type_override: str
    startup_bias: str
    notes: str


@dataclass(frozen=True)
class ResumeOutreachQueueItem:
    row_id: str
    company: str
    role_title: str
    status: str
    date_found: date | None
    fit_score: float | None
    outreach_priority_score: float
    company_type: str
    startup_bias: str
    priority_reasons: list[str]
    source: str
    source_url: str
    url_hash: str


def load_resume_jobs(path: Path, sheet_name: str = "Jobs") -> list[ResumeJob]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for importing ResumeGenerator jobs.xlsx."
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
        sheet = workbook[sheet_name]
        headers = [_clean_cell(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows: list[ResumeJob] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            raw = {headers[index]: _clean_cell(value) for index, value in enumerate(values)}
            if not any(raw.values()):
                continue
            company = raw.get("company", "").strip()
            role_title = raw.get("role_title", "").strip()
            url = raw.get("url", "").strip()
            if not company or not role_title:
                continue
            status = raw.get("status", "").strip()
            rows.append(
                ResumeJob(
                    row_id=raw.get("id", "").strip(),
                    company=company,
                    role_title=role_title,
                    location=raw.get("location", "").strip(),
                    url=url,
                    url_hash=raw.get("url_hash", "").strip(),
                    source=raw.get("source", "").strip(),
                    status=status,
                    normalized_status=normalize_resume_status(status),
                    fit_score=_parse_float(raw.get("fit_score", "")),
                    fit_rationale=raw.get("fit_rationale", "").strip(),
                    date_found=_parse_date(raw.get("date_found", "")),
                    date_posted_raw=raw.get("date_posted", "").strip(),
                    folder_path=raw.get("folder_path", "").strip(),
                    jd_text=raw.get("jd_text", "").strip(),
                    notes=raw.get("notes", "").strip(),
                )
            )
        return rows
    finally:
        workbook.close()


def select_resume_jobs(
    jobs: Iterable[ResumeJob],
    *,
    include_statuses: Iterable[str] = DEFAULT_INCLUDE_STATUSES,
    min_score: float = 7.0,
    max_age_days: int | None = 10,
    today: date | None = None,
) -> ResumeImportSelection:
    filtered: list[ResumeJob] = []
    skipped_missing_identity = 0
    skipped_status = 0
    skipped_score = 0
    skipped_age = 0
    duplicates_removed = 0
    normalized_requested = tuple(include_statuses)
    allowed_statuses = {
        normalize_resume_status(item) for item in (normalized_requested or DEFAULT_INCLUDE_STATUSES)
    }
    reference_day = today or date.today()

    for job in jobs:
        if not job.company or not job.role_title:
            skipped_missing_identity += 1
            continue
        if job.normalized_status not in allowed_statuses:
            skipped_status += 1
            continue
        if job.fit_score is None or job.fit_score < min_score:
            skipped_score += 1
            continue
        if max_age_days is not None:
            if job.date_found is None:
                skipped_age += 1
                continue
            if (reference_day - job.date_found).days > max_age_days:
                skipped_age += 1
                continue
        filtered.append(job)

    selected, duplicates_removed = dedupe_resume_jobs(filtered)
    return ResumeImportSelection(
        jobs=selected,
        skipped_missing_identity=skipped_missing_identity,
        skipped_status=skipped_status,
        skipped_score=skipped_score,
        skipped_age=skipped_age,
        duplicates_removed=duplicates_removed,
    )


def normalize_resume_status(value: str) -> str:
    return value.strip().lower()


def dedupe_resume_jobs(jobs: Iterable[ResumeJob]) -> tuple[list[ResumeJob], int]:
    by_key: dict[str, ResumeJob] = {}
    duplicates_removed = 0
    for job in jobs:
        dedupe_key = dedupe_key_for_job(job)
        existing = by_key.get(dedupe_key)
        if existing is None:
            by_key[dedupe_key] = job
            continue
        duplicates_removed += 1
        if resume_job_sort_key(job) > resume_job_sort_key(existing):
            by_key[dedupe_key] = job

    selected = sorted(by_key.values(), key=resume_job_sort_key, reverse=True)
    return selected, duplicates_removed


def dedupe_key_for_job(job: ResumeJob) -> str:
    if job.url_hash:
        return f"url_hash:{job.url_hash.strip().lower()}"
    company = normalize_dedupe_text(job.company)
    role = normalize_dedupe_text(job.role_title)
    return f"company_role:{company}|{role}"


def normalize_dedupe_text(value: str) -> str:
    lowered = value.strip().lower()
    return re.sub(r"\s+", " ", lowered)


def resume_job_sort_key(job: ResumeJob) -> tuple[date, float, str, str, str]:
    return (
        job.date_found or date.min,
        job.fit_score if job.fit_score is not None else -1.0,
        normalize_resume_status(job.status),
        normalize_dedupe_text(job.company),
        normalize_dedupe_text(job.role_title),
    )


def map_resume_source_kind(source: str) -> SourceKind:
    lowered = source.strip().lower()
    if lowered in {"linkedin", "linkedin_live_jobs_v1", "screenshot"}:
        return SourceKind.LINKEDIN_JOB
    if lowered in {"indeed", "seeded"}:
        return SourceKind.JOB_BOARD
    return SourceKind.OTHER


def infer_opportunity_type(role_title: str) -> OpportunityType:
    lowered = role_title.lower()
    if any(keyword in lowered for keyword in ("intern", "internship", "co-op", "coop")):
        return OpportunityType.INTERNSHIP
    if any(keyword in lowered for keyword in ("new grad", "graduate")):
        return OpportunityType.FULL_TIME
    return OpportunityType.OTHER


def opportunity_status_from_resume_status(status: str) -> str:
    normalized = normalize_resume_status(status)
    mapping = {
        "queued": "outreach_ready",
        "generated": "assets_ready",
        "applied": "applied_catchup",
    }
    return mapping.get(normalized, "imported")


def organization_status_from_resume_status(status: str) -> str:
    normalized = normalize_resume_status(status)
    if normalized in {"queued", "generated"}:
        return "Pre-apply outreach"
    if normalized == "applied":
        return "Applied outreach catch-up"
    return "Imported"


def target_lists_from_resume_status(status: str) -> str:
    normalized = normalize_resume_status(status)
    if normalized == "applied":
        return "jobs;resume_generator;post_apply_catchup"
    return DEFAULT_TARGET_LISTS


def ensure_company_overrides_csv(path: Path) -> Path:
    if path.exists():
        return path
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "company,company_type_override,startup_bias,notes\n",
        encoding="utf-8",
    )
    return path


def load_company_overrides(path: Path) -> dict[str, CompanyOverride]:
    import csv

    ensure_company_overrides_csv(path)
    overrides: dict[str, CompanyOverride] = {}
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            company = str(row.get("company") or "").strip()
            if not company:
                continue
            normalized_company = normalize_dedupe_text(company)
            overrides[normalized_company] = CompanyOverride(
                company=company,
                normalized_company=normalized_company,
                company_type_override=normalize_company_type(str(row.get("company_type_override") or "")),
                startup_bias=normalize_startup_bias(str(row.get("startup_bias") or "")),
                notes=str(row.get("notes") or "").strip(),
            )
    return overrides


def normalize_company_type(value: str) -> str:
    normalized = normalize_dedupe_text(value).replace(" ", "_")
    if normalized in {"startup", "growth", "big_company", "company"}:
        return normalized
    return "unknown"


def normalize_startup_bias(value: str) -> str:
    normalized = normalize_dedupe_text(value)
    if normalized in STARTUP_BIAS_SCORES:
        return normalized
    return "neutral"


def infer_company_type_for_job(
    job: ResumeJob,
    *,
    company_override: CompanyOverride | None = None,
) -> str:
    if company_override and company_override.company_type_override != "unknown":
        return company_override.company_type_override
    source = normalize_dedupe_text(job.source)
    company = normalize_dedupe_text(job.company)
    if source == "screenshot":
        return "unknown"
    big_company_signals = {
        "tiktok",
        "ibm",
        "cisco",
        "nvidia",
        "comcast",
        "gen",
    }
    growth_signals = {
        "typeface",
        "cosm",
        "arena",
    }
    if company in big_company_signals:
        return "big_company"
    if company in growth_signals:
        return "growth"
    return "unknown"


def organization_type_for_resume_job(
    job: ResumeJob,
    *,
    company_override: CompanyOverride | None = None,
) -> OrganizationType:
    company_type = infer_company_type_for_job(job, company_override=company_override)
    if company_type in {"startup", "growth"}:
        return OrganizationType.STARTUP
    return OrganizationType.COMPANY


def build_resume_outreach_queue(
    jobs: Iterable[ResumeJob],
    *,
    company_overrides: dict[str, CompanyOverride] | None = None,
    max_per_company: int = 2,
) -> list[ResumeOutreachQueueItem]:
    company_overrides = company_overrides or {}
    per_company_counts: dict[str, int] = {}
    items: list[ResumeOutreachQueueItem] = []
    for job in sorted(jobs, key=resume_job_sort_key, reverse=True):
        normalized_company = normalize_dedupe_text(job.company)
        if per_company_counts.get(normalized_company, 0) >= max_per_company:
            continue
        override = company_overrides.get(normalized_company)
        company_type = infer_company_type_for_job(job, company_override=override)
        startup_bias = override.startup_bias if override else "neutral"
        priority_score, reasons = compute_outreach_priority(
            job,
            company_type=company_type,
            startup_bias=startup_bias,
        )
        items.append(
            ResumeOutreachQueueItem(
                row_id=job.row_id,
                company=job.company,
                role_title=job.role_title,
                status=job.normalized_status,
                date_found=job.date_found,
                fit_score=job.fit_score,
                outreach_priority_score=priority_score,
                company_type=company_type,
                startup_bias=startup_bias,
                priority_reasons=reasons,
                source=job.source,
                source_url=job.url,
                url_hash=job.url_hash,
            )
        )
        per_company_counts[normalized_company] = per_company_counts.get(normalized_company, 0) + 1

    items.sort(
        key=lambda item: (
            item.outreach_priority_score,
            item.date_found or date.min,
            item.fit_score if item.fit_score is not None else -1.0,
            normalize_dedupe_text(item.company),
        ),
        reverse=True,
    )
    return items


def compute_outreach_priority(
    job: ResumeJob,
    *,
    company_type: str,
    startup_bias: str,
) -> tuple[float, list[str]]:
    score = job.fit_score if job.fit_score is not None else 0.0
    reasons = [f"fit_score={score:.1f}"]
    company_bonus = COMPANY_TYPE_SCORES.get(company_type, 0.0)
    if company_bonus:
        reasons.append(f"company_type={company_type}")
        score += company_bonus
    bias_bonus = STARTUP_BIAS_SCORES.get(startup_bias, 0.0)
    if bias_bonus:
        reasons.append(f"startup_bias={startup_bias}")
        score += bias_bonus
    if job.normalized_status == "generated":
        reasons.append("assets_ready")
        score += 0.4
    if job.normalized_status == "applied":
        reasons.append("applied_catchup")
    freshness_bonus = freshness_bonus_for_job(job)
    if freshness_bonus:
        reasons.append("fresh")
        score += freshness_bonus
    return round(score, 2), reasons


def freshness_bonus_for_job(job: ResumeJob) -> float:
    if job.date_found is None:
        return 0.0
    age_days = (date.today() - job.date_found).days
    if age_days <= 3:
        return 1.0
    if age_days <= 7:
        return 0.5
    return 0.0


def build_resume_opportunity_notes(job: ResumeJob) -> str:
    parts = [
        f"resume_job_id={job.row_id}",
        f"resume_status={job.normalized_status or 'unknown'}",
        f"resume_source={job.source or 'unknown'}",
    ]
    if job.url_hash:
        parts.append(f"url_hash={job.url_hash}")
    if job.fit_score is not None:
        parts.append(f"fit_score={job.fit_score:.1f}")
    if job.date_found is not None:
        parts.append(f"date_found={job.date_found.isoformat()}")
    if job.folder_path:
        parts.append(f"folder_path={job.folder_path}")
    if job.fit_rationale:
        parts.append(f"fit_rationale={job.fit_rationale}")
    return " | ".join(parts)


def build_resume_organization_notes(job: ResumeJob) -> str:
    parts = [
        "Imported from ResumeGenerator v1 jobs.xlsx",
        f"latest_resume_status={job.normalized_status or 'unknown'}",
    ]
    if job.date_found is not None:
        parts.append(f"latest_date_found={job.date_found.isoformat()}")
    return " | ".join(parts)


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _parse_float(value: str) -> float | None:
    text = value.strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_date(value: str) -> date | None:
    text = value.strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None
