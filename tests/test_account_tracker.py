from pathlib import Path

from openpyxl import load_workbook

from outreach.account_tracker import (
    DailyPlanBudget,
    audit_track_2_core,
    build_account_rows,
    build_campaign_plan_rows,
    build_track_2_daily_plan,
    generate_excel,
)
from outreach.tracking import ContactRecord, OpportunityRecord, OrganizationRecord, OrganizationType, OutreachWorkbook


def test_account_tracker_filters_unrelated_existing_connections(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-clara",
            name="Clara",
            organization_type=OrganizationType.STARTUP,
            notes=(
                "team_size=9 | tags=healthcare,ai | "
                "description=Clara is an AI primary care assistant."
            ),
        )
    )
    workbook.upsert_contact(
        ContactRecord(
            contact_id="ct-founder",
            organization_id="org-clara",
            full_name="George Founder",
            title="Founder at Clara",
            status="Discovered",
        )
    )
    workbook.upsert_contact(
        ContactRecord(
            contact_id="ct-warm-unrelated",
            organization_id="org-clara",
            full_name="Warm Unrelated",
            title="Senior Principal Software Engineer at Optum",
            status="Warm",
            notes="passes=existing_connections | triggers=Existing Connection,USC Marshall",
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.company == "Clara"
    assert row.people_mapped == 1
    assert row.accepted == 0
    assert row.replies == 0
    assert row.score_relationship == 0
    assert row.account_stage == "people_mapped"


def test_account_tracker_counts_company_relevant_warm_contact_as_accepted(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-centerfield",
            name="Centerfield",
            organization_type=OrganizationType.COMPANY,
            notes="tags=data-platform,analytics | description=Marketing technology and data platform.",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-centerfield",
            organization_id="org-centerfield",
            title="Product Manager Intern",
            opportunity_type="internship",
        )
    )
    workbook.upsert_contact(
        ContactRecord(
            contact_id="ct-product",
            organization_id="org-centerfield",
            full_name="Product Person",
            title="Product Manager at Centerfield",
            status="Warm",
            notes="passes=existing_connections | triggers=Existing Connection",
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.people_mapped == 1
    assert row.accepted == 1
    assert row.replies == 0
    assert row.score_relationship == 3
    assert row.score_relationship_momentum == 8
    assert row.account_stage == "connected_no_conversation"


def test_account_priority_is_not_dominated_by_relationship_momentum(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-vercel",
            name="Vercel",
            organization_type=OrganizationType.COMPANY,
            notes=(
                "tags=developer-tools,platform-engineering,data-platform,artificial-intelligence | "
                "description=Frontend cloud, AI developer tooling, and data workflow platform."
            ),
        )
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-warm",
            name="WarmOps",
            organization_type=OrganizationType.COMPANY,
            notes=(
                "tags=developer-tools,platform-engineering,data-platform,artificial-intelligence | "
                "description=Developer workflow tools, AI automation, and data platform workflows for internal teams."
            ),
        )
    )
    workbook.upsert_contact(
        ContactRecord(
            contact_id="ct-warm",
            organization_id="org-warm",
            full_name="Warm Contact",
            title="Product Manager at WarmOps",
            status="Replied",
        )
    )

    rows = build_account_rows(tmp_path)
    by_company = {row.company: row for row in rows}
    plan = build_campaign_plan_rows(rows)

    assert by_company["Vercel"].account_score > by_company["WarmOps"].account_score
    assert by_company["WarmOps"].daily_action_priority > by_company["Vercel"].daily_action_priority
    assert plan[0].company == "WarmOps"


def test_account_tracker_parses_team_size_with_commas_and_words(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-deel",
            name="Deel",
            organization_type=OrganizationType.COMPANY,
            notes="team_size=5,000 employees | tags=fintech,hr-tech | description=Global payroll platform.",
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.team_size == 5000
    assert row.score_team_gate == 0
    assert "team_size_unparsed" not in row.data_quality_flags


def test_account_tracker_profile_fit_uses_token_boundaries(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-capital-fulfillment",
            name="Capital Fulfillment",
            organization_type=OrganizationType.COMPANY,
            notes=(
                "team_size=225 Employees | tags=fintech,financial services | "
                "description=Capital Fulfillment hires talented teams for high trust member support."
            ),
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.score_profile_fit == 4
    assert "AI/ML" not in row.why_fit
    assert "LLM/AI" not in row.why_fit
    assert "API/integration" not in row.why_fit
    assert "talent" not in row.why_fit


def test_account_tracker_manual_priority_counts_without_stacking_brand(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-dream-data",
            name="Dream Data",
            organization_type=OrganizationType.COMPANY,
            target_lists="relationship;dream",
            notes="tags=data-platform | description=Data platform for enterprise workflows.",
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.score_brand == 12
    assert "manual priority" in row.why_fit
    assert row.account_score > row.fit_score


def test_account_tracker_story_fit_metadata_drives_pitch_strength(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-anam",
            name="Anam AI",
            organization_type=OrganizationType.COMPANY,
            target_lists="story-fit;track-2;relationship",
            notes=(
                "source=story_fit_targets | tags=artificial-intelligence,workflow-automation | "
                "why_this_company=FlairX gives a direct recruiting workflow pitch. | "
                "story_angle=hiring interview AI | profile_evidence=FlairX AI PM internship."
            ),
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.score_pitch_strength >= 8
    assert "explicit story-fit pitch" in row.why_fit
    assert "profile evidence" in row.why_fit


def test_account_tracker_prestige_signals_feed_brand_component(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-funded-ai",
            name="Funded AI",
            organization_type=OrganizationType.COMPANY,
            notes=(
                "tags=artificial-intelligence | description=AI workflow platform. | "
                "prestige_signals=sequoia-backed,series-b | context_confidence=external_verified"
            ),
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.score_brand == 10
    assert "top investor signal" in row.why_fit


def test_account_campaign_plan_switches_channel_after_large_dead_linkedin_wave(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-centerfield",
            name="Centerfield",
            organization_type=OrganizationType.COMPANY,
            notes="tags=data-platform,analytics | description=Marketing technology and data platform.",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-centerfield",
            organization_id="org-centerfield",
            title="Product Manager Intern",
            opportunity_type="internship",
        )
    )
    for index in range(9):
        workbook.upsert_contact(
            ContactRecord(
                contact_id=f"ct-{index}",
                organization_id="org-centerfield",
                full_name=f"Person {index}",
                title=f"Product Manager at Centerfield {index}",
                status="Invited",
            )
        )

    row = build_account_rows(tmp_path)[0]

    assert row.account_stage == "outreach_active"
    assert row.campaign_action == "find_email_path"
    assert row.campaign_channel == "email_research"
    assert row.lane_1_policy == "track_2_owns"


def test_account_campaign_plan_uses_email_after_dead_linkedin_wave_when_email_exists(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-centerfield",
            name="Centerfield",
            organization_type=OrganizationType.COMPANY,
            notes="tags=data-platform,analytics | description=Marketing technology and data platform.",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-centerfield",
            organization_id="org-centerfield",
            title="Product Manager Intern",
            opportunity_type="internship",
        )
    )
    for index in range(9):
        workbook.upsert_contact(
            ContactRecord(
                contact_id=f"ct-{index}",
                organization_id="org-centerfield",
                full_name=f"Person {index}",
                title=f"Product Manager at Centerfield {index}",
                email="product@example.com" if index == 0 else "",
                status="Invited",
            )
        )

    row = build_account_rows(tmp_path)[0]

    assert row.email_contacts == 1
    assert row.campaign_action == "send_cold_email_followup"
    assert row.campaign_channel == "email"
    assert row.lane_1_policy == "track_2_owns"


def test_account_campaign_plan_allows_parallel_initial_outreach_when_email_exists(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-agent-startup",
            name="Agent Startup",
            organization_type=OrganizationType.STARTUP,
            target_lists="story-fit;track-2;relationship;priority",
            notes=(
                "team_size=40 | tags=artificial-intelligence,data-platform,workflow-automation | "
                "description=AI workflow platform for enterprise data operations."
            ),
        )
    )
    workbook.upsert_contact(
        ContactRecord(
            contact_id="ct-founder",
            organization_id="org-agent-startup",
            full_name="Founder Person",
            title="Founder at Agent Startup",
            email="founder@example.com",
            status="Discovered",
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.tier == "A"
    assert row.email_contacts == 1
    assert row.campaign_action == "send_initial_multichannel_outreach"
    assert row.campaign_channel == "linkedin+email"


def test_track_2_core_audit_is_clean_for_generated_multichannel_row(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-agent-startup",
            name="Agent Startup",
            organization_type=OrganizationType.STARTUP,
            target_lists="story-fit;track-2;relationship;priority",
            notes=(
                "team_size=40 | tags=artificial-intelligence,data-platform,workflow-automation | "
                "description=AI workflow platform for enterprise data operations."
            ),
        )
    )
    workbook.upsert_contact(
        ContactRecord(
            contact_id="ct-founder",
            organization_id="org-agent-startup",
            full_name="Founder Person",
            title="Founder at Agent Startup",
            email="founder@example.com",
            status="Discovered",
        )
    )

    audit = audit_track_2_core(build_account_rows(tmp_path))

    assert audit["is_clean"] is True
    assert audit["issue_counts"] == {}


def test_track_2_daily_plan_enforces_budgets(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    for index in range(4):
        workbook.upsert_organization(
            OrganizationRecord(
                organization_id=f"org-follow-{index}",
                name=f"Follow {index}",
                organization_type=OrganizationType.STARTUP,
                notes=(
                    "team_size=40 | tags=artificial-intelligence,data-platform | "
                    "description=AI data platform."
                ),
            )
        )
        workbook.upsert_contact(
            ContactRecord(
                contact_id=f"ct-follow-{index}",
                organization_id=f"org-follow-{index}",
                full_name=f"Warm Person {index}",
                title=f"Product Manager at Follow {index}",
                status="Warm",
            )
        )
    for index in range(4):
        workbook.upsert_organization(
            OrganizationRecord(
                organization_id=f"org-map-{index}",
                name=f"Map {index}",
                organization_type=OrganizationType.STARTUP,
                target_lists="story-fit;track-2;relationship;priority",
                notes=(
                    "team_size=40 | tags=artificial-intelligence,data-platform,workflow-automation | "
                    "description=AI workflow platform."
                ),
            )
        )

    plan = build_track_2_daily_plan(
        build_account_rows(tmp_path),
        budget=DailyPlanBudget(
            max_total_actions=5,
            max_companies=5,
            max_linkedin_invites=0,
            max_linkedin_followups=2,
            max_company_mapping=2,
            max_email_research=1,
            max_context_enrichment=0,
            max_email_drafts=0,
        ),
    )

    assert plan["used"]["total_actions"] <= 5
    assert plan["used"]["linkedin_followups"] <= 2
    assert plan["used"]["company_mapping"] <= 2
    assert plan["used"]["email_research"] <= 1
    assert plan["skipped_count"] > 0
    selected = plan["selected"]
    assert selected == sorted(
        selected,
        key=lambda item: (item["phase_order"], -item["daily_action_priority"], item["company"].lower()),
    )
    assert plan["phase_summary"]
    assert plan["execution_order"][0]["phase"] == "1_continue_live_conversations"
    mapping_items = [item for item in selected if item["campaign_action"] == "map_more_contacts"]
    assert mapping_items
    assert all(item["phase"] == "4_contact_mapping" for item in mapping_items)
    assert any(item["expected_email_research"] == 1 for item in mapping_items)


def test_track_2_daily_plan_degrades_mapping_to_linkedin_only_when_email_research_capped(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-story-fit",
            name="StoryFit",
            organization_type=OrganizationType.STARTUP,
            target_lists="story-fit;track-2;relationship;priority;tier-a",
            notes=(
                "team_size=40 | tags=artificial-intelligence,data-platform,workflow-automation | "
                "description=AI workflow platform for recruiting teams. | "
                "why_this_company=Direct story fit | profile_evidence=Engineering plus MBA | "
                "story_angle=AI workflow/productivity"
            ),
        )
    )

    plan = build_track_2_daily_plan(
        build_account_rows(tmp_path),
        budget=DailyPlanBudget(
            max_total_actions=2,
            max_companies=2,
            max_linkedin_invites=0,
            max_linkedin_followups=0,
            max_company_mapping=2,
            max_email_research=0,
            max_context_enrichment=0,
            max_email_drafts=0,
        ),
    )

    item = plan["selected"][0]
    assert item["campaign_action"] == "map_more_contacts"
    assert item["campaign_channel"] == "linkedin"
    assert item["expected_company_mapping"] == 1
    assert item["expected_email_research"] == 0
    assert plan["used"]["company_mapping"] == 1
    assert plan["used"]["email_research"] == 0


def test_identity_conflict_account_is_paused(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-conflict",
            name="ConflictCo",
            organization_type=OrganizationType.COMPANY,
            target_lists="story-fit;track-2;relationship;priority;tier-a",
            notes=(
                "identity_conflict=true | tags=artificial-intelligence,data-platform | "
                "description=Conflicting company source data."
            ),
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert "identity_conflict" in row.data_quality_flags
    assert row.campaign_action == "pause_account"
    assert row.campaign_channel == "none"
    assert build_campaign_plan_rows([row]) == []


def test_account_campaign_plan_flags_role_without_domain_context(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-typeface",
            name="Typeface",
            organization_type=OrganizationType.COMPANY,
            notes="Imported from ResumeGenerator v1 jobs.xlsx | latest_resume_status=generated",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-typeface",
            organization_id="org-typeface",
            title="Product Manager Intern",
            opportunity_type="internship",
        )
    )

    row = build_account_rows(tmp_path)[0]
    plan = build_campaign_plan_rows([row])[0]

    assert "needs_domain_enrichment" in row.data_quality_flags
    assert row.account_score < row.fit_score
    assert row.campaign_action == "enrich_company_context"
    assert plan.campaign_action == "enrich_company_context"
    assert plan.account_score == row.account_score
    assert plan.lane_1_policy == "fresh_role_only"


def test_account_hiring_prefers_ft_product_path_over_summer_internship(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    shared_notes = (
        "team_size=120 | tags=artificial-intelligence,data-platform,developer-tools | "
        "description=AI data platform for developer workflows."
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-summer",
            name="SummerOnly",
            organization_type=OrganizationType.COMPANY,
            notes=shared_notes,
        )
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-ft",
            name="FullTimePath",
            organization_type=OrganizationType.COMPANY,
            notes=shared_notes,
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-summer",
            organization_id="org-summer",
            title="Product Manager Intern (Summer 2026)",
            opportunity_type="internship",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-ft",
            organization_id="org-ft",
            title="Associate Product Manager",
            opportunity_type="full_time",
        )
    )

    by_company = {row.company: row for row in build_account_rows(tmp_path)}

    assert by_company["FullTimePath"].score_hiring > by_company["SummerOnly"].score_hiring
    assert by_company["FullTimePath"].score_account_hiring > by_company["SummerOnly"].score_account_hiring
    assert by_company["FullTimePath"].account_score > by_company["SummerOnly"].account_score
    assert by_company["SummerOnly"].hiring_signal == "Summer internship discovered"


def test_account_hiring_counts_la_or_remote_fall_internship_more_than_summer(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    shared_notes = (
        "team_size=90 | tags=artificial-intelligence,data-platform | "
        "description=AI data workflow platform."
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-fall",
            name="FallPath",
            organization_type=OrganizationType.STARTUP,
            notes=shared_notes,
        )
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-summer",
            name="SummerPath",
            organization_type=OrganizationType.STARTUP,
            notes=shared_notes,
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-fall",
            organization_id="org-fall",
            title="Product Manager Intern - Fall 2026",
            opportunity_type="internship",
            location="Los Angeles, CA (Hybrid)",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-summer",
            organization_id="org-summer",
            title="Product Manager Intern - Summer 2026",
            opportunity_type="internship",
        )
    )

    by_company = {row.company: row for row in build_account_rows(tmp_path)}

    assert by_company["FallPath"].score_hiring > by_company["SummerPath"].score_hiring
    assert by_company["FallPath"].score_account_hiring > by_company["SummerPath"].score_account_hiring
    assert by_company["FallPath"].hiring_signal == "Fall/co-op internship, LA/remote"


def test_account_hiring_discounts_in_person_fall_internship_outside_la(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    shared_notes = (
        "team_size=90 | tags=artificial-intelligence,data-platform | "
        "description=AI data workflow platform."
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-remote",
            name="RemoteFall",
            organization_type=OrganizationType.STARTUP,
            notes=shared_notes,
        )
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-ny",
            name="NewYorkFall",
            organization_type=OrganizationType.STARTUP,
            notes=shared_notes,
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-remote",
            organization_id="org-remote",
            title="Product Management Intern, Fall 2026",
            opportunity_type="internship",
            location="United States (Remote)",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-ny",
            organization_id="org-ny",
            title="Product Management Intern (Fall 2026)",
            opportunity_type="internship",
            location="New York, NY (Hybrid)",
        )
    )

    by_company = {row.company: row for row in build_account_rows(tmp_path)}

    assert by_company["RemoteFall"].score_hiring == 12
    assert by_company["NewYorkFall"].score_hiring == 1
    assert by_company["RemoteFall"].score_account_hiring > by_company["NewYorkFall"].score_account_hiring
    assert by_company["NewYorkFall"].hiring_signal == "Fall/co-op internship outside LA/remote"


def test_account_hiring_season_detection_ignores_location_substrings(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-buffalo",
            name="Buffalo Role",
            organization_type=OrganizationType.STARTUP,
            notes=(
                "team_size=90 | tags=artificial-intelligence,data-platform | "
                "description=AI data workflow platform."
            ),
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-buffalo",
            organization_id="org-buffalo",
            title="Product Owner Internship",
            opportunity_type="internship",
            location="Buffalo-Niagara Falls Area",
        )
    )

    row = build_account_rows(tmp_path)[0]

    assert row.score_hiring == 6
    assert row.hiring_signal == "Internship discovered"


def test_relationship_tier_a_excludes_large_companies(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()

    for index in range(22):
        workbook.upsert_organization(
            OrganizationRecord(
                organization_id=f"org-startup-{index}",
                name=f"Startup {index:02d}",
                organization_type=OrganizationType.STARTUP,
                target_lists="yc;startup",
                notes=(
                    "team_size=50 | tags=artificial-intelligence,data-platform,developer-tools | "
                    "description=AI developer data workflow platform."
                ),
            )
        )
    for index in range(12):
        workbook.upsert_organization(
            OrganizationRecord(
                organization_id=f"org-growth-{index}",
                name=f"Growth {index:02d}",
                organization_type=OrganizationType.COMPANY,
                notes=(
                    "team_size=450 | tags=artificial-intelligence,data-platform,developer-tools | "
                    "description=AI developer data workflow platform."
                ),
            )
        )
    for index in range(12):
        workbook.upsert_organization(
            OrganizationRecord(
                organization_id=f"org-large-{index}",
                name=f"Large {index:02d}",
                organization_type=OrganizationType.COMPANY,
                target_lists="dream",
                notes=(
                    "team_size=5000 | tags=artificial-intelligence,data-platform,developer-tools | "
                    "description=AI developer data workflow platform."
                ),
            )
        )

    rows = build_account_rows(tmp_path)
    tier_a = [row for row in rows if row.tier == "A"]
    large_rows = [row for row in rows if row.account_track == "Large Company"]
    track_counts = {}
    for row in tier_a:
        track_counts[row.account_track] = track_counts.get(row.account_track, 0) + 1

    assert len(tier_a) == 32
    assert "Large Company" not in track_counts
    assert track_counts["Startup / Founder-Led"] >= 20
    assert track_counts["Growth / Mid-Market"] >= 12
    assert {row.tier for row in large_rows} == {"L1"}


def test_account_tracker_excel_writes_operational_views(tmp_path: Path) -> None:
    workbook = OutreachWorkbook(tmp_path)
    workbook.initialize()
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-startup",
            name="Agent Startup",
            organization_type=OrganizationType.STARTUP,
            target_lists="yc;startup",
            notes="team_size=18 | tags=artificial-intelligence | description=AI agent platform.",
        )
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-growth",
            name="Growth Data",
            organization_type=OrganizationType.COMPANY,
            notes="team_size=450 employees | tags=data-platform | description=Enterprise data platform.",
        )
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-large",
            name="BigCo",
            organization_type=OrganizationType.COMPANY,
            notes="team_size=3,500 employees | tags=developer-tools | description=Developer tools platform.",
        )
    )
    workbook.upsert_organization(
        OrganizationRecord(
            organization_id="org-needs-context",
            name="Thin Import",
            organization_type=OrganizationType.COMPANY,
            notes="Imported from ResumeGenerator v1 jobs.xlsx | latest_resume_status=generated",
        )
    )
    workbook.upsert_opportunity(
        OpportunityRecord(
            opportunity_id="opp-thin",
            organization_id="org-needs-context",
            title="Product Manager Intern",
            opportunity_type="internship",
        )
    )

    rows = build_account_rows(tmp_path)
    output = generate_excel(rows, tmp_path / "account_tracker.xlsx")
    wb = load_workbook(output, read_only=True)

    assert "Startup Founder-Led" in wb.sheetnames
    assert "Growth Mid-Market" in wb.sheetnames
    assert "Large Company" in wb.sheetnames
    assert "Large Company Priority" in wb.sheetnames
    assert "Strategic Wishlist" in wb.sheetnames
    assert "Needs Enrichment" in wb.sheetnames

    by_company = {row.company: row for row in rows}
    assert by_company["Agent Startup"].account_track == "Startup / Founder-Led"
    assert by_company["Growth Data"].account_track == "Growth / Mid-Market"
    assert by_company["BigCo"].account_track == "Large Company"
    assert by_company["Thin Import"].campaign_action == "enrich_company_context"
